import datetime
import requests
import luigi
import openpyxl
import csv
from luigi import Task, LocalTarget
import luigi.contrib.spark
from pyspark.sql import SparkSession
from prometheus_client import start_http_server, Summary
import time
import os
from luigi.contrib.spark import PySparkTask

values_time = Summary('request_processing_seconds', 'Time spent processing request')
spark_time = Summary('request_sparkprocessing_seconds', 'Time spent sparkprocessing request')

class DownloadData(Task):
    filename = 'data.xlsx'
    api = 'https://cloud-api.yandex.net/v1/disk/public/resources/download?public_key={}'
    def directlink(self):
        pk_request = requests.get(self.api.format('https://yadi.sk/i/awTwxgB2pXn8xw'))
        return pk_request.json().get('href')

    def run(self):

        direct_link = self.directlink()
        if direct_link:
            download = requests.get(direct_link)
            with open(self.filename, 'wb') as f:
                f.write(download.content)

    def output(self):
        return LocalTarget(self.filename)


class TimeSelect(Task):
    filename = "timecoord.txt"

    def requires(self):
        return DownloadData()

    def run(self):
        bf = openpyxl.load_workbook('data.xlsx')
        wl = bf.active
        allb = wl['B']
        tod = datetime.datetime.now()
        delta = datetime.timedelta(days=11)
        fulltime = tod - delta
        f = open(self.filename, 'w')
        for cell in allb:
            if type(cell.value) == type(fulltime):
                if cell.value > fulltime:
                    c = cell.coordinate[1:]
                    f.write(c + '\n')

    def output(self):
        return LocalTarget(self.filename)


class FiledSelect(Task):
    filename = "filedcoord.txt"

    def requires(self):
        return TimeSelect()

    def run(self):
        bf = openpyxl.load_workbook('data.xlsx')
        wl = bf.active
        alle = wl['E']
        f = open(self.filename, 'w')
        with self.input().open('r') as cdate:
            for coord in cdate:
                if alle[int(coord)].value is not None:
                    c = alle[int(coord)].coordinate[1:]
                    f.write(c + '\n')

    def output(self):
        return LocalTarget(self.filename)


class ValidValues(Task):
    filename = "valueslist.csv"

    def requires(self):
        return FiledSelect()

    @values_time.time()
    def run(self):
        bf = openpyxl.load_workbook('data.xlsx')
        wl = bf.active
        valuerows = []
        listrows = []
        tod = datetime.datetime.now()
        with self.input().open('r') as valide:
            for line in valide:
                cur = line[:-1]
                listrows.append(cur)
        validcells = wl['B' + listrows[0]:'G' + listrows[-1]]
        for row in validcells:
            for cell in row:
                if 'B' in cell.coordinate:
                    valuerows.append(cell.value.strftime('%m/%d/%Y'))
                elif 'C' in cell.coordinate:
                    valuerows.append(cell.value.strftime('%H:%M'))
                else:
                    valuerows.append(cell.value)
        rowslist = [valuerows[i:i + 6] for i in range(0, len(valuerows), 6)]
        writer = csv.writer(open(self.filename, 'w', newline='', encoding='UTF-8'), delimiter=',')
        writer.writerow(['Дата', 'Время', 'Номер', 'Производитель', 'Источник', 'Штрафстоянка'])
        for line in rowslist:
            writer.writerow(line)

    def output(self):
        return LocalTarget(self.filename)

'''
###Fails on Windows :(

class SparkCount(luigi.contrib.spark.SparkSubmitTask):
    app = "sparkproc.py"

    def output(self):
        return LocalTarget("countstreet")# next

    def requires(self):
        return ValidValues()

    def app_options(self):
       return [self.input().path]
'''
'''
class SparkCount(Task):

    def output(self):
        return LocalTarget("countstreet")  # next

    def requires(self):
        return ValidValues()

    @spark_time.time()
    def run(self):
        input = self.input().path

        spark = SparkSession.builder \
            .config("spark.driver.bindAddress", "127.0.0.1") \
            .master("local[*]") \
            .appName("Row Count") \
            .getOrCreate()

        df = spark.read.csv(input, header="true")
        hotstreet = df.groupBy("source").count()

        hotstreet.coalesce(1).write.csv("countstreet", sep=',', encoding='UTF-8', header='True')

        spark.stop()
'''


if __name__ == '__main__':
    start_http_server(8000)
    luigi.build([ValidValues()])
    #luigi.run()
    while True:
        pass
